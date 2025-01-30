import pandas as pd
import jinja2
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import os

def style_border(file_name):

    "This function will set boarders to the cells of excel"
    wb = load_workbook(file_name)
    ws = wb.active

    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter  # Get the column letter (e.g., A, B, C)
        for cell in column_cells:
            try:
                if cell.value:  # Check if the cell has a value
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_name)


def PosCash():

    '''This function will process epost file'''
    file_path =  str(input("Enter file_path: "))
    data = pd.read_excel(file_path)
    cols = ['Profit Center', 'Profit Centre Text',
        'Posting Date', 'Closing Balance']
    data = data[cols]
    data = data[data['Closing Balance'] != 0]
    data.sort_values('Posting Date', inplace = True)
    y_day = date.today()-timedelta(days=1)
    data['Posting Date'] = data['Posting Date'].apply(lambda x: x.date())
    data = data[data['Posting Date'] < y_day]
    data['Posting Date'] = data['Posting Date'].apply(lambda x: x.strftime('%d-%m-%Y'))
    data.style.map(lambda x: 'color: red;')
    file_name = 'PosCash'
    file_path = 'Reports'
    write_file(data, file_path, file_name)


def write_file(data, file_path, file_name):

    """ This function will write files to excel """
    file_path = os.path.join(file_path, file_name)
    file_name = file_path+'.xlsx'
    data.to_excel(file_name, index = False)
    style_border(file_name)


def eMO():

    """This function will generate eMO file"""
    file_path =  str(input("Enter file_path: "))
    data = pd.read_excel(file_path)
    cols = ['Office Name','Not Printed Unpaid Emos',
        'Printed Unpaid Emos', 'Total Unpaid Emos']
    data = data[cols]
    #return data, 'eMO'
    file_name = 'eMO'
    file_path = 'Reports'
    write_file(data, file_path, file_name)

def ePost():

    """This function will generate ePost file"""
    file_path =  str(input("Enter file_path: "))
    data = pd.read_excel(file_path, skiprows = 8)
    cols = ['S.No.', 'ePost Center', 'Retail', 'Prepaid', 'Corporate', 'Total']
    data = data.iloc[:, [2,11,12,13,16]]
    data.rename(columns = {'Unnamed: 1':'S.No.', 'Unnamed: 2':'ePost Center',
                        'Retail .1':'Retail', 'Prepaid .1':'Prepaid', 'Corporate .1':'Corporate',
                        'Total.1': 'Total'}, inplace = True)
    data = data[data.Total>0]
    data.iloc[-1,0] = 'Total'
    #return data, 'ePost'
    file_name = 'ePost'
    file_path = 'Reports'
    write_file(data, file_path, file_name)
    